"""
SCRIPT DE CARGA DE DATOS DESDE EXCEL A POSTGRESQL

Este script toma datos de un archivo Excel y los carga a una base de datos PostgreSQL.
Maneja tres hojas específicas: clientes, productos y ventas, en ese orden.

Cómo usar:
1. Asegurarse de tener el archivo Excel en la ruta correcta
2. Configura la conexión a la DB en config/database.py
3. Ejecuta el script: python cargar_datos.py

Características:
- Verifica que el archivo y hojas existan antes de procesar
- Carga los datos manteniendo los existentes (no borra información previa)
- Proporciona feedback claro durante el proceso
- Maneja errores de forma elegante

CARGA DE DATOS - CONFIGURACIÓN EXTERNA SIMPLE

Llamamos siempre la configuracion de la base de datos desde un archivo aparte 
para no tener las credenciales visibles en el codigo principal, y en el
.gitignore bloqueamos las vistas.

"""

import pandas as pd
from sqlalchemy import create_engine, text
import os
from openpyxl import load_workbook
import sys
from pathlib import Path

# Configuración de rutas
# (Permite importar módulos desde la carpeta config)
sys.path.append(str(Path(__file__).parent.parent.parent))
from config.database import get_db_uri

# Ruta al archivo Excel con los datos
EXCEL_PATH = "./data/input/datos_fuente.xlsx"

def verificar_archivo_excel():
    """
    Revisa que el archivo Excel esté en su lugar y tenga las hojas necesarias.
    
    Si algo falla, detiene el programa con un mensaje claro de qué salió mal.
    """
    # Verificar si el archivo existe
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"❌ No encuentro el archivo Excel en: {EXCEL_PATH}")
    
    # Cargar el libro de trabajo y verificar hojas
    wb = load_workbook(EXCEL_PATH)
    hojas_necesarias = ['clientes', 'productos', 'ventas']
    hojas_en_excel = wb.sheetnames
    
    # Verificar cada hoja requerida
    for hoja in hojas_necesarias:
        if hoja not in hojas_en_excel:
            raise ValueError(
                f"❌ La hoja '{hoja}' no está en el Excel. "
                f"Hojas encontradas: {', '.join(hojas_en_excel)}"
            )
    
    print("✓ Archivo Excel verificado correctamente")
    return True

def cargar_datos(engine, hoja):
    """
    Carga los datos de una hoja Excel a la base de datos.
    
    Args:
        engine: Conexión a la base de datos
        hoja: Nombre de la hoja a cargar (clientes, productos o ventas)
    
    Returns:
        bool: True si la carga fue exitosa, False si hubo error
    """
    try:
        print(f"\nProcesando hoja: {hoja}...")
        
        # Leer datos del Excel
        datos = pd.read_excel(EXCEL_PATH, sheet_name=hoja)
        
        # Verificar si hay datos
        if datos.empty:
            print(f"⚠️ Advertencia: La hoja '{hoja}' está vacía")
            return False
        
        # Cargar a PostgreSQL
        datos.to_sql(
            name=hoja,
            con=engine,
            if_exists="append",  # Añade sin borrar datos existentes
            index=False
        )
        
        print(f"✓ Datos cargados: {len(datos)} registros en '{hoja}'")
        return True
        
    except Exception as error:
        print(f"❌ Error al cargar '{hoja}': {str(error)}")
        return False

def main():
    """
    Flujo principal del programa:
    1. Verifica el archivo Excel
    2. Conecta a la base de datos
    3. Carga los datos en orden
    4. Proporciona un resumen final
    """
    print("\n" + "="*50)
    print("  INICIANDO CARGA DE DATOS EXCEL → POSTGRESQL")
    print("="*50)
    
    try:
        # PASO 1: Verificar que el Excel esté correcto
        verificar_archivo_excel()
        
        # PASO 2: Conectar a la base de datos
        print("\nConectando a la base de datos...")
        engine = create_engine(get_db_uri())
        print("✓ Conexión establecida")
        
        # PASO 3: Cargar datos en el orden correcto
        # (primero clientes, luego productos, finalmente ventas)
        hojas_a_cargar = ['clientes', 'productos', 'ventas']
        resultados = []
        
        for hoja in hojas_a_cargar:
            resultados.append(cargar_datos(engine, hoja))
        
        # PASO 4: Mostrar resumen final
        print("\n" + "="*50)
        if all(resultados):
            print("✅ CARGA COMPLETADA: Todos los datos se cargaron correctamente")
        else:
            print("⚠️ CARGA PARCIAL: Algunos datos no se cargaron (ver errores arriba)")
        print("="*50)
            
    except Exception as error:
        print("\n" + "❌"*10)
        print(f"ERROR CRÍTICO: {str(error)}")
        print("❌"*10)
    finally:
        # Asegurarse de cerrar la conexión a la DB
        if 'engine' in locals():
            engine.dispose()
            print("\nConexión a la base de datos cerrada")

# Punto de entrada del script
if __name__ == "__main__":
    main()